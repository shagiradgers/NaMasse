# Для работы нужна библиотека PyQt5_stylesheets
# pip install git+https://github.com/RedFalsh/PyQt5_stylesheets

import sys
from PyQt5.QtWidgets import QApplication, QWidget, \
    QTableWidgetItem, QDialog, QInputDialog, QFileDialog, QMessageBox
from PyQt5 import uic
import sqlite3
from time import asctime
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
import xlsxwriter
import PyQt5_stylesheets

# Таблица с данными
DATA_NAME = 'some data.sqlite'


class Window(QWidget):
    def __init__(self):
        global DATA_NAME
        super(Window, self).__init__()
        uic.loadUi('design_to_tz.ui', self)
        self.con = sqlite3.connect(DATA_NAME)
        self.message = QMessageBox(self)
        self.main_data = []
        self.row = 0
        self.get_data()
        self.table.setEditTriggers(self.table.NoEditTriggers)
        self.btn_enter_data.clicked.connect(self.btn_pressed)
        self.btn_save_xls.clicked.connect(self.btn_pressed)
        self.btn_graphic.clicked.connect(self.btn_pressed)
        self.btn_rec.clicked.connect(self.btn_pressed)
        self.btn_settings.clicked.connect(self.btn_pressed)
        self.table.cellDoubleClicked.connect(self.table_clicked)

    # обработчик нажатий на кнопку
    def btn_pressed(self):
        if self.sender() == self.btn_enter_data:
            dialog.show()

        elif self.sender() == self.btn_save_xls:
            if len(self.recommendation(self.main_data)) > 1:
                self.save_xls()
            else:
                self.message.setText('Недостаточно данных')
                self.message.exec()

        elif self.sender() == self.btn_graphic:
            if len(self.recommendation(self.main_data)) > 1:
                self.draw_graphic(self.main_data)
            else:
                self.message.setText('Недостаточно данных')
                self.message.exec()
        elif self.sender() == self.btn_settings:
            settings.show()

        elif self.sender() == self.btn_rec:
            if len(self.recommendation(self.main_data)) > 1:
                dialog_rec.update_data()
                dialog_rec.show()
            else:
                self.message.setText('Недостаточно данных')
                self.message.exec()

    # Обработчик нажатий на таблицу
    def table_clicked(self, row):
        answer, ok_pressed = QInputDialog.getItem(
            self, "Вопрос", "Что вы хотите сделать?",
            ("Удалить строчку", "Изменить входные данные"), 1, False)
        if answer == 'Удалить строчку' and ok_pressed:
            self.del_item(row)

        elif answer == 'Изменить входные данные' and ok_pressed:
            self.row = row
            dialog_calculation.show()

    # Удаление строки
    def del_item(self, row):
        self.table.removeRow(row)
        rows = self.table.rowCount()
        cols = self.table.columnCount()
        data = []
        for row in range(rows):
            tmp = []
            for col in range(cols):
                try:
                    tmp.append(int(self.table.item(row, col).text()))
                except ValueError:
                    tmp.append(self.table.item(row, col).text())
            data.append(tmp)
        self.main_data = data
        self.draw_table(data)

        cur = self.con.cursor()
        cur.execute('''DELETE FROM calculated_data''')
        self.con.commit()
        cur.execute('''DELETE FROM user_data''')
        self.con.commit()
        cur.execute('''DELETE FROM age''')
        self.con.commit()
        cur.execute('''DELETE FROM time''')
        self.con.commit()

    # Замена строки
    def change_item(self, row, data):
        cur = self.con.cursor()
        cur.execute('''UPDATE calculated_data SET body_mass_index = ?, 
        calories = ?, protein = ?, fats = ?, carbohydrates = ? 
        WHERE id_user = ?''', (round(data['body_mass_index']),
                               round(data['calories']),
                               round(data['protein']),
                               round(data['fats']),
                               round(data['carbohydrates']),
                               row)).fetchall()
        self.con.commit()
        cur.execute('''UPDATE user_data SET weight = ?, 
        height = ?, calories = ?, protein = ?, fats = ?, 
        carbohydrates = ? WHERE id_user = ?''',
                    (data['weight'],
                     data['height'],
                     round(data['user_calories']),
                     round(data['user_protein']),
                     round(data['user_fats']),
                     round(data['user_carbohydrates']),
                     row)).fetchall()
        self.con.commit()
        cur.execute('''UPDATE time SET time = ?,
         id_time = ?''', (data['time'],
                          row)).fetchall()
        self.con.commit()
        cur.execute('''UPDATE age SET age = ?,
        id_age = ?''', (data['age'],
                        row)).fetchall()
        self.con.commit()

        self.get_data()

    # отрисовка таблицы
    def draw_table(self, data):
        title = ['Индекс массы тела', 'Рекомендуемые калории',
                 'Рекомендуемые белки', 'Рекомендуемые жиры',
                 'Рекомендуемые углеводы', 'Вес', 'Рост',
                 'Калории', 'Белки', 'Жиры', 'Углеводы',
                 'Время', 'Возраст']
        self.table.setColumnCount(len(title))
        self.table.setHorizontalHeaderLabels(title)
        self.table.setRowCount(0)
        for i, row in enumerate(data):
            self.table.setRowCount(
                self.table.rowCount() + 1)
            for j, elem in enumerate(row):
                self.table.setItem(
                    i, j, QTableWidgetItem(str(elem)))
        self.table.resizeColumnsToContents()

    # добаваление элементов в self.table
    def add_to_table(self, data):
        cur = self.con.cursor()
        cur.execute('''INSERT INTO 
        calculated_data(body_mass_index, calories, 
        protein, fats, carbohydrates, id_user)
        VALUES(?, ?, ?, ?, ?, ?)''', (round(data['body_mass_index']),
                                      round(data['calories']),
                                      round(data['protein']),
                                      round(data['fats']),
                                      round(data['carbohydrates']),
                                      len(self.main_data))).fetchall()
        self.con.commit()
        cur.execute('''INSERT INTO
        user_data(weight, height, calories, protein,
         fats, carbohydrates,id_time, id_age ,id_user)
        VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                    (data['weight'],
                     data['height'],
                     round(data['user_calories']),
                     round(data['user_protein']),
                     round(data['user_fats']),
                     round(data['user_carbohydrates']),
                     len(self.main_data),
                     len(self.main_data),
                     len(self.main_data))).fetchall()
        self.con.commit()
        cur.execute('''INSERT INTO time(
        time, id_time)
        VALUES(?, ?)''', (data['time'], len(self.main_data)))
        self.con.commit()
        cur.execute('''INSERT INTO age(
        age, id_age)
        VALUES(?, ?)''', (data['age'], len(self.main_data)))
        self.con.commit()
        self.get_data()

    # получение данных из sql таблицы
    def get_data(self):
        cur = self.con.cursor()
        try:

            data_calculated = cur.execute('''SELECT body_mass_index, 
            calories, protein, fats, carbohydrates, id_user 
            FROM calculated_data''').fetchall()
            data_user = cur.execute('''SELECT weight, height, 
            calories, protein, fats, carbohydrates, id_time,
             id_age, id_user 
            FROM user_data''').fetchall()
            data_time = cur.execute('''SELECT time, id_time 
            FROM time''').fetchall()
            data_age = cur.execute('''SELECT age, id_age
            FROM age''').fetchall()
            data = []

            # соединение data_calculated, data_user, data_time и data_age
            for calculated_data in data_calculated:
                for user_data in data_user:
                    for time in data_time:
                        for age in data_age:
                            if calculated_data[-1] == user_data[-1]:
                                if user_data[-3] == time[-1]:
                                    if user_data[-2] == age[-1]:
                                        data.append(calculated_data[:-1] +
                                                    user_data[:-3] +
                                                    tuple([time[0]]) +
                                                    tuple([str(age[0])]))
                                        break

        except sqlite3.OperationalError:
            cur.execute("""CREATE TABLE IF NOT EXISTS calculated_data(
            body_mass_index INTEGER,
            calories REAL,
            protein REAL,
            fats REAL,
            carbohydrates INTEGER,
            id_user INTEGER PRIMARY KEY AUTOINCREMENT
                    NOT NULL
                    REFERENCES user_data (id_user));""")
            self.con.commit()
            cur.execute('''CREATE TABLE IF NOT EXISTS user_data(
            weight INTEGER,
            height INTEGER,
            calories REAL,
            protein REAL,
            fats REAL,
            carbohydrates REAL,
            id_time INTEGER REFERENCES time (id_time),
            id_age  INTEGER REFERENCES age (id_age),
            id_user INTEGER PRIMARY KEY AUTOINCREMENT
                      NOT NULL
                      REFERENCES user_data (id_user));''')
            self.con.commit()
            cur.execute('''CREATE TABLE time(
            time    TEXT,
            id_time INTEGER PRIMARY KEY AUTOINCREMENT
                    NOT NULL
                    REFERENCES user_data (id_time));''')
            self.con.commit()
            cur.execute('''CREATE TABLE IF NOT EXISTS age(
            age    INTEGER,
            id_age INTEGER PRIMARY KEY AUTOINCREMENT
                   NOT NULL
                   REFERENCES user_data (id_age));''')
            self.con.commit()
            data = []
        self.main_data = data
        self.draw_table(data)

    # сохранение таблицы в xlsx
    def save_xls(self):
        file_name, ok = QFileDialog.getSaveFileName(self,
                                                    "Сохранить файл",
                                                    "YourResults.xlsx",
                                                    "*xlsx")
        # если расширение не xlsx
        if file_name.split('/')[-1].split('.')[-1] != 'xlsx':
            file_name = '/'.join(file_name.split('/')[:-1] + [
                '.'.join([file_name.split('/')[-1].split('.')[0], 'xlsx'])])

        workbook = xlsxwriter.Workbook(file_name)
        worksheet = workbook.add_worksheet()
        cur = self.con.cursor()
        data = [['Индекс массы тела', 'Рекомендуемые калории',
                 'Рекомендуемые белки', 'Рекомендуемые жиры',
                 'Рекомендуемые углеводы', 'Вес', 'Рост',
                 'Калории', 'Белки', 'Жиры', 'Углеводы',
                 'Время', 'Возраст']]

        data_calculated = cur.execute('''SELECT body_mass_index, 
                    calories, protein, fats, carbohydrates, id_user 
                    FROM calculated_data''').fetchall()
        data_user = cur.execute('''SELECT weight, height, 
                    calories, protein, fats, carbohydrates,
                     id_time ,id_age, id_user 
                    FROM user_data''').fetchall()
        data_time = cur.execute('''SELECT time, id_time
                                FROM time''').fetchall()
        data_age = cur.execute('''SELECT age, id_age
                               FROM age''').fetchall()

        # соединение data_calculated, data_user, data_time и data_age
        for calculated_data in data_calculated:
            for user_data in data_user:
                for time in data_time:
                    for age in data_age:
                        if calculated_data[-1] == user_data[-1]:
                            if user_data[-3] == time[-1]:
                                if user_data[-2] == age[-1]:
                                    data.append(calculated_data[:-1] +
                                                user_data[:-3] +
                                                tuple([time[0]]) +
                                                tuple([age[0]]))
                                    break

        for element_1, row in enumerate(data):
            for element_2, column in enumerate(row):
                worksheet.write(element_1, element_2, column)

        workbook.close()

    # Построение графика
    def draw_graphic(self, some_data):
        file_name, ok = QFileDialog.getSaveFileName(self,
                                                    "Сохранить файл",
                                                    "Graphic.xlsx",
                                                    "*xlsx")

        # если расширение не xlsx
        if file_name.split('/')[-1].split('.')[-1] != 'xlsx':
            file_name = '/'.join(file_name.split('/')[:-1] + [
                '.'.join([file_name.split('/')[-1].split('.')[0], 'xlsx'])])

        wb = Workbook(write_only=True)
        wb.create_sheet(title='Графики', index=0)
        wb.create_sheet(title='Калории', index=1)
        wb.create_sheet(title='Белки', index=2)
        wb.create_sheet(title='Жиры', index=3)
        wb.create_sheet(title='Углеводы', index=4)

        main_sheet = wb['Графики']
        sheet_cal = wb['Калории']
        sheet_protein = wb['Белки']
        sheet_fats = wb['Жиры']
        sheet_carb = wb['Углеводы']

        batch1 = [('Date',
                   'Рекомендуемое количество калорий',
                   'Ваше количество калорий')]
        batch2 = [('Date',
                   'Рекомендуемое количество белков',
                   'Ваше количесвто белков')]
        batch3 = [('Date',
                   'Рекомендуемое количество жиров',
                   'Ваше количество жиров')]
        batch4 = [('Date',
                   'Рекомендуемое количество углеводов',
                   'Ваше количество углеводов')]

        for element in some_data:
            batch1.append((element[-2], element[1], element[-6]))
            batch2.append((element[-2], element[2], element[-5]))
            batch3.append((element[-2], element[3], element[-4]))
            batch4.append((element[-2], element[4], element[-3]))

        max_row = len(batch1)

        for row in batch1:
            sheet_cal.append(row)
        for row in batch2:
            sheet_protein.append(row)
        for row in batch3:
            sheet_fats.append(row)
        for row in batch4:
            sheet_carb.append(row)

        chart1 = BarChart()
        chart1.type = "col"
        chart1.style = 10
        chart1.title = "Калории"
        chart1.y_axis.title = 'Количество'
        chart1.x_axis.title = 'Дата'

        data_cal = Reference(sheet_cal,
                             min_col=2,
                             min_row=1,
                             max_row=max_row,
                             max_col=3)
        cats_cal = Reference(sheet_cal,
                             min_col=1,
                             min_row=2,
                             max_row=max_row)
        chart1.add_data(data_cal,
                        titles_from_data=True)
        chart1.set_categories(cats_cal)
        chart1.shape = 4
        main_sheet.add_chart(chart1, "A1")

        chart2 = BarChart()
        chart2.type = "col"
        chart2.style = 10
        chart2.title = "Белки"
        chart2.y_axis.title = 'Количесвто'
        chart2.x_axis.title = 'Дата'
        data_protein = Reference(sheet_protein,
                                 min_col=2,
                                 min_row=1,
                                 max_row=max_row,
                                 max_col=3)
        cats_protein = Reference(sheet_protein,
                                 min_col=1,
                                 min_row=2,
                                 max_row=max_row)
        chart2.add_data(data_protein,
                        titles_from_data=True)
        chart2.set_categories(cats_protein)
        chart2.shape = 4
        main_sheet.add_chart(chart2, "I1")

        chart3 = BarChart()
        chart3.type = "col"
        chart3.style = 10
        chart3.title = "Жиры"
        chart3.y_axis.title = 'Количество'
        chart3.x_axis.title = 'Дата'
        data_fats = Reference(sheet_fats,
                              min_col=2,
                              min_row=1,
                              max_row=max_row,
                              max_col=3)
        cats_fats = Reference(sheet_fats,
                              min_col=1,
                              min_row=2,
                              max_row=max_row)
        chart3.add_data(data_fats,
                        titles_from_data=True)
        chart3.set_categories(cats_fats)
        chart3.shape = 4
        main_sheet.add_chart(chart3, "A15")

        chart4 = BarChart()
        chart4.type = "col"
        chart4.style = 10
        chart4.title = "Углеводы"
        chart4.y_axis.title = 'Количесвто'
        chart4.x_axis.title = 'Дата'
        data_carb = Reference(sheet_carb,
                              min_col=2,
                              min_row=1,
                              max_row=max_row,
                              max_col=3)
        cats_carb = Reference(sheet_carb,
                              min_col=1,
                              min_row=2,
                              max_row=max_row)
        chart4.add_data(data_carb,
                        titles_from_data=True)
        chart4.set_categories(cats_carb)
        chart4.shape = 4
        main_sheet.add_chart(chart4, "I15")

        wb.save(file_name)

    # получение рекомендаций для user'a
    @staticmethod
    def recommendation(data):
        list_of_recommendation = []
        # data[...][1] - рекомендуемое количество калорий
        # data[...][-6] - калории user'a
        try:
            if int(data[-1][1]) < int(data[-1][-6]):
                list_of_recommendation.append('''
                Рекомендуем снизить количество
                                          потребляемых калорий на
                                          {}'''.format(round(
                    int(data[-1][-6]) - int(data[-1][1]))))
            elif int(data[-1][1]) > int(data[-1][-6]):
                list_of_recommendation.append('''Рекомендуем потреблять на
                                          {} калорий больше,
                                           чем сейчас'''.format(
                    round(int(data[-1][1]) - int(data[-1][-6]))))
            else:
                list_of_recommendation.append('''Вы потребляете достаточное 
                                            количество калорий''')

            # data[...][3] - рекомендуемое количество белков
            # data[...][-5] - белки user'a

            if int(data[-1][-5]) > int(data[-1][3]):
                list_of_recommendation.append('''
                Рекомендуем снизить количество
                                            потреблямых 
                                            белков на {} г.'''.format(
                    int(data[-1][-5]) - int(data[-1][3])))
            elif int(data[-1][-5]) < int(data[-1][3]):
                list_of_recommendation.append('''
                Рекомендуем увеличть количество
                                            потребляемых 
                                            белков на {} г.'''.format(int(
                    round(data[-1][3]) - int(data[-1][-5]))))
            else:
                list_of_recommendation.append('''Вы потребляете достаточное 
                количество белков''')

            # data[...][3] - рекомендуемое количество жиров
            # data[...][-4] - жиры user'a

            if int(data[-1][-4]) > int(data[-1][3]):
                list_of_recommendation.append('''
                Рекомендуем снизить количество
                              потрбления жиров на {}  г.'''.format(
                    round(int(
                        data[-1][-4]) - int(data[-1][3]))))
            elif int(data[-1][-4]) < int(data[-1][3]):
                list_of_recommendation.append('''
                Рекомендуем увеличить количество
                              потребления жиров на {} г.'''.format(
                    round(int(
                        data[-1][3]) - int(data[-1][-4]))))
            else:
                list_of_recommendation.append('''Вы потребляете достаточное 
            количество жиров''')
            # data[...][4] - рекомендуемое количество углеводов
            # data[...][-3] - углеводы user'a

            if int(data[-1][-3]) > int(data[-1][4]):
                list_of_recommendation.append('''
                Рекомендуем снизить количество
                              потребляемых углеводов на {} г.'''.format(int(
                    round(data[-1][-3] - int(data[-1][4])))))
            elif int(data[-1][-3]) < int(data[-1][4]):
                list_of_recommendation.append('''
                Рекомендуем увеличить количество
                              потребляемых углеводов на {} г.'''.format(int(
                    round(data[-1][4]) - int(data[-1][-3]))))
            else:
                list_of_recommendation.append('''Вы потребляете достаточное 
            количество углеводов''')
            some_list = []
            for some_recommendation in list_of_recommendation:
                some_list.append([some_data for some_data in
                                  some_recommendation.split()
                                  if some_data != '\n' or some_data != ''])

            return [' '.join(some_data) for some_data in some_list]
        except IndexError:
            return []


class Calculation(QDialog):
    def __init__(self, mode):
        super(Calculation, self).__init__()
        uic.loadUi('dialog_to_design.ui', self)
        self.mode = mode
        self.dialogButtonBox.accepted.connect(self.accepted)

    # если user нажал 'ok' в диалоговом окне
    def accepted(self):
        try:
            if all([data.split() != [] and int(data) for data in
                    [self.weight_data.text(), self.height_data.text(),
                     self.age_data.text(), self.cal_data.text(),
                     self.protein_data.text(), self.fats_data.text(),
                     self.carb_data.text()]]):
                self.calculate({'weight': int(self.weight_data.text()),
                                'height': int(self.height_data.text()),
                                'age': int(self.age_data.text()),
                                'calories': int(self.cal_data.text()),
                                'protein': int(self.protein_data.text()),
                                'fats': int(self.fats_data.text()),
                                'carbohydrates': int(self.carb_data.text())})
                app.log_zone.setText('')
            else:
                app.log_zone.setText('Неправильный ввод')
        except ValueError:
            app.log_zone.setText('Неправильный ввод')

    # функция для подсчета
    def calculate(self, data):
        # Коэффициент полезного действия
        coefficient = 0
        if self.combobox_activity.currentText() == 'минимальная активность':
            coefficient = 1.2
        elif self.combobox_activity.currentText() == 'умеренная активность':
            coefficient = 1.55
        elif self.combobox_activity.currentText() == 'высокая активность':
            coefficient = 1.725
        elif self.combobox_activity.currentText() == 'экстра активность':
            coefficient = 1.9

        user_fats = data['fats']
        user_calories = data['calories']
        user_protein = data['protein']
        user_carbohydrates = data['carbohydrates']

        body_mass_index = round(float(data['weight']) / (
                (float(data['height']) / 100) ** 2))

        # Формула подсчета Миффлин-Сан Жеора
        # (10 x вес (кг) + 6.25 x рост (см) –
        # 5 x возраст (г) + 5) x A - для мужчин.
        # (10 x вес (кг) + 6.25 x рост (см) –
        # 5 x возраст (г) - 161) x A - для женщин.
        if self.combobox_gender.currentText() == 'Мужской':
            calories = round((10 * float(data['weight']) + 6.25 * float(
                data['height']) - 5 * float(data['age']) + 5
                              ) * coefficient)
        else:
            calories = round((10 * float(data['weight']) + 6.25 * float(
                data['height']) - 5 * float(data['age']) - 161
                              ) * coefficient)
        if self.combobox_goal.currentText() == 'Поддерживать вес':
            protein = round(((calories - calories * 0.15)
                             * 0.3) / 4)
            fats = round(((calories - calories * 0.15)
                          * 0.3) / 9)
            carbohydrates = round(((calories - calories * 0.15)
                                   * 0.4) / 4)
        else:
            protein = ((calories - calories * 0.15)
                       * 0.3) / 4
            fats = ((calories - calories * 0.15)
                    * 0.2) / 9
            carbohydrates = ((calories - calories * 0.15)
                             * 0.5) / 4

        if self.mode == 1:
            app.add_to_table({'body_mass_index': body_mass_index,
                              'calories': calories,
                              'protein': protein,
                              'fats': fats,
                              'carbohydrates': carbohydrates,
                              'time': asctime(),
                              'weight': data['weight'],
                              'height': data['height'],
                              'age': data['age'],
                              'user_calories': user_calories,
                              'user_protein': user_protein,
                              'user_fats': user_fats,
                              'user_carbohydrates': user_carbohydrates})
        else:
            app.change_item(app.row, {'body_mass_index': body_mass_index,
                                      'calories': calories,
                                      'protein': protein,
                                      'fats': fats,
                                      'carbohydrates': carbohydrates,
                                      'time': asctime(),
                                      'weight': data['weight'],
                                      'height': data['height'],
                                      'age': data['age'],
                                      'user_calories': user_calories,
                                      'user_protein': user_protein,
                                      'user_fats': user_fats,
                                      'user_carbohydrates':
                                          user_carbohydrates})


class Preview(QDialog):
    def __init__(self):
        super(Preview, self).__init__()
        uic.loadUi('preview.ui', self)


class Faq(QDialog):
    def __init__(self):
        super(Faq, self).__init__()
        uic.loadUi('FAQ.ui', self)


class Recommendation(QDialog):
    def __init__(self):
        super(Recommendation, self).__init__()
        uic.loadUi('recommendation_design.ui', self)
        list_of_recommendation = app.recommendation(app.main_data)
        for data in list_of_recommendation:
            self.list_recommendation.addItem(data)

    # обновление данных
    def update_data(self):
        self.list_recommendation.clear()
        list_of_recommendation = app.recommendation(app.main_data)
        for data in list_of_recommendation:
            self.list_recommendation.addItem(data)


class Settings(QWidget):
    def __init__(self):
        super(Settings, self).__init__()
        uic.loadUi('settings.ui', self)
        self.btn_faq.clicked.connect(self.pressed)
        self.btn_preview.clicked.connect(self.pressed)
        self.btn_theme.clicked.connect(self.pressed)

    # обработчик нажатий на кнопку
    def pressed(self):
        if self.sender() == self.btn_faq:
            dialog_faq.show()
        elif self.sender() == self.btn_preview:
            dialog_preview.show()
        else:
            answer, ok_pressed = QInputDialog.getItem(
                self, "Вопрос", "Какую тему вы хотите использовать?",
                ("Светло-синию", "Светло-темную", "Классическую тему",
                 "Темную тему", "Темно-орнажевую", "Светло-серую"), 1, False)
            if answer == 'Светло-синию' and ok_pressed:
                self.change_style('style_blue')
            elif answer == 'Светло-темную' and ok_pressed:
                self.change_style('style_black')
            elif answer == 'Классическую тему' and ok_pressed:
                self.change_style('style_Classic')
            elif answer == 'Темную тему' and ok_pressed:
                self.change_style('style_Dark')
            elif answer == 'Темно-орнажевую' and ok_pressed:
                self.change_style('style_DarkOrange')
            elif answer == 'Светло-серую' and ok_pressed:
                self.change_style('style_gray')

    # смена цветовой схемы приложения
    @staticmethod
    def change_style(style):
        # смена цветового оформления программы
        app.setStyleSheet(
            PyQt5_stylesheets.load_stylesheet_pyqt5(
                style=style
            )
        )
        app.message.setStyleSheet(
            PyQt5_stylesheets.load_stylesheet_pyqt5(
                style=style
            )
        )
        dialog_faq.setStyleSheet(
            PyQt5_stylesheets.load_stylesheet_pyqt5(
                style=style
            )
        )
        dialog_rec.setStyleSheet(
            PyQt5_stylesheets.load_stylesheet_pyqt5(
                style=style
            )
        )
        dialog_preview.setStyleSheet(
            PyQt5_stylesheets.load_stylesheet_pyqt5(
                style=style
            )
        )
        dialog_calculation.setStyleSheet(
            PyQt5_stylesheets.load_stylesheet_pyqt5(
                style=style
            )
        )
        dialog.setStyleSheet(
            PyQt5_stylesheets.load_stylesheet_pyqt5(
                style=style
            )
        )
        settings.setStyleSheet(
            PyQt5_stylesheets.load_stylesheet_pyqt5(
                style=style
            )
        )


if __name__ == '__main__':
    wnd = QApplication(sys.argv)
    app = Window()
    app.show()
    dialog = Calculation(1)
    dialog_calculation = Calculation(0)
    dialog_preview = Preview()
    dialog_faq = Faq()
    dialog_rec = Recommendation()
    settings = Settings()
    sys.exit(wnd.exec_())
